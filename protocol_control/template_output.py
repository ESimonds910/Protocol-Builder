# from docx import Document
# from docx.text.paragraph import Paragraph
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
# from docx.opc.coreprops import CoreProperties
# import docx2txt


class Templates:
    def __init__(self, template_choice):
        self.choice = template_choice
        # self.pick_templates()

    def pick_templates(self):
        if self.choice == 'Spaniel':
            cell_dict = {
                'Project_Name': 'B4',
                'Date': 'C1',
                'SourcePlates': 'D9',
                'GreinerPlates': 'D10',
                'ProxiPlates': 'D11',
                'Standard_conc': ['B18', 'B19', 'B20', 'B21', 'B22', 'B23'],
                'Volumes': ['D57', 'D58', 'D59', 'D60'],
                'data': 32
            }
            wb = load_workbook(r"L:\High Throughput Screening\Personnel\Matthew Currie\Standard Protocol_SSF.xlsx")
            ws = wb['Protocol']
            return cell_dict, wb, ws


class TemplateOutput:
    def __init__(self, output_dict):
        self.cells, self.wb, self.ws = Templates(output_dict['Project_Name']).pick_templates()
        self.output_dict = output_dict
        self.output_folder = r"L:\High Throughput Screening\Personnel\Matthew Currie"
        self.output_file = self.output_folder + r'\test_output.xlsx'

        self.excel_wb()
        self.df = pd.DataFrame()

    def excel_wb(self):

        for key, value in self.output_dict.items():
            if isinstance(value, list):
                mini_dict = dict(zip(self.cells[key], value))
                for mini_key, mini_value in mini_dict.items():
                    self.ws[mini_key].value = mini_value

            elif isinstance(value, pd.DataFrame):
                start_row = self.cells[key]
                for r in dataframe_to_rows(value, index=False, header=True):
                    start_col = 2
                    for c in r:
                        self.ws.cell(row=start_row, column=start_col).value = c
                        start_col += 1
                    start_row += 1
            else:
                try:
                    self.ws[self.cells[key]].value = value
                except KeyError:
                    pass

        cell_to_append = self.ws['B27:B29']


        self.wb.save(self.output_file)

        # self.ws.append("B27", test_list)
        # for num in test_list:
        #     self.ws.cell(row=27, column=test_list.index(num) + 2).value = num
        # self.ws[f'B27:B{27 + len(test_list)}'] = test_list
        # self.ws["B27"] = 'test'
        # self.ws.move_range("B27", cols=1)
        # n = 0
        # for value in self.ws.iter_rows(min_row=27, max_row=27, min_col=2, max_col=len(test_list)):
        #     value = test_list[n]
        #     n += 1
        # for r in dataframe_to_rows(self.output_dict['data'], index=True, header=True):
        #     self.ws.append(r)
        # for n in range(5):
        #     self.ws.insert_rows(27)

    def read_doc(self):
        with open(self.file, encoding='utf-16') as doc_to_read:
            for line in doc_to_read:
                print(line)

    def find_keywords(self):
        # keywords = self.template.core_properties
        # for keyword in keywords:
        #     print(keyword.keywords)
        body = self.template._body
        # assuming differentiating container element is w:textBox
        text_box_p_elements = body.xpath('.//w:textBox//w:p')

    def find_sections(self):
        paragraphs = self.template.paragraphs
        # for key, value in self.output_dict.items():
        #     print(key)
        #     for paragraph in paragraphs:
        #         split_paragraph = paragraph.text.split(' ')
        #         for word in split_paragraph:
        #             if key == word:
        #                 word = word.replace(key, value)
        #                 paragraph.text = ' '.join(split_paragraph)
        for key, value in self.output_dict.items():
            for p in paragraphs:
                print(p.text)
                inline = p.runs
                print(inline)
                # for i in range(len(inline)):
                #     text = inline[i].text
                #     print(text)
                #     if key in text:
                #         text = text.replace(key, value)
                #         inline[i].text = text

                # print(word.text)
        # if '[Project_Name]' in paragraphs.text:
        #     paragraph.text = paragraph.text.replace('[Project_Name]', self.project)
        # print(paragraph.text)

        # self.template.save(r'L:\High Throughput Screening\Personnel\Matthew Currie\text.docx')
    def print_outputs(self):
        print(self.project)
        # print(CoreProperties(self.template).keywords)
        print(Paragraph(self.template).text)
        # for line in self.template[:2]:
        #     print(line)


if __name__ == "__main__":
    test_df = pd.DataFrame([[1, 2, 3], [4, 5, 6]], columns=["A", "B", "C"])
    test_dict = {
        'Project_Name': 'Spaniel',
        'SourcePlates': 8,
        'GreinerPlates': 8,
        'ProxiPlates': 10,
        'Date': '8/4/2021',
        'data': test_df,
        'Standard_conc': [30, 10, 3.3, 1.1, 0.4, 0.14]
    }
    TemplateOutput(test_dict)
